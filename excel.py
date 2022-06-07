# Reading an excel file using Python

import pandas as pd

import os

import openpyxl
 
# Give the location of the file
path = "C:\\Users\\masis\\water.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

# Dictionary to hold the date value pairs
my_dict = {"Date":[],"Value":[]};
 
# number of coulmns of excel file
max_col = sheet_obj.max_column
# number of rows of excel file
m_row = sheet_obj.max_row
for i in range(2, m_row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 1)
    x=cell_obj.value.split()
    line=x[12].split(";")
    date=line[4]+" "+ line[5]
    height=line[7]
    
    my_dict["Date"].append(date)
    my_dict["Value"].append(height)
for i in range(len(my_dict["Date"])):    
 print ("Date"+ my_dict["Date"][i] +" = " + my_dict["Value"][i])   
    
 